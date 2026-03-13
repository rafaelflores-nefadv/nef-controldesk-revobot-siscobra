from datetime import datetime
from decimal import Decimal, InvalidOperation
import logging
import socket
from pathlib import Path
import re
import shutil
from time import sleep
import os
import ssl
from ftplib import FTP, FTP_TLS, error_perm
import pandas as pd
from openpyxl import Workbook

from core.download_api import (
    baixar_arquivo_via_form_submit_no_browser,
    listar_arquivos_api_no_browser,
)

from config.settings import (
    FILE_MANAGER_EXPORT_FOLDER,
    DOWNLOAD_DIR,
    DOWNLOAD_WAIT_POLL_SECONDS,
    DOWNLOAD_WAIT_TIMEOUT_SECONDS,
    FILE_PREFIX,
    COPY_DIR,
    FTP_HOST,
    FTP_USER,
    FTP_PASSWORD,
    FTP_DIR,
    FTP_USE_TLS,
    FTP_PORT,
    FTP_TLS_IMPLICIT,
    FTP_TLS_VERIFY,
    FTP_TIMEOUT,
    FTP_TLS_LEGACY,
    FTP_DATA_PROTECTION,
    FTP_PASSIVE,
)

logger = logging.getLogger(__name__)


def _normalize_cpf_cnpj(value: str) -> str:
    value = str(value).strip()
    if not value:
        return ""

    normalized = value
    if "e" in value.lower():
        scientific_value = value.replace(",", ".")
        try:
            normalized = str(int(Decimal(scientific_value)))
        except (InvalidOperation, ValueError, OverflowError):
            normalized = value

    digits = re.sub(r"\D+", "", normalized)
    if not digits:
        return ""
    if len(digits) <= 11:
        return digits.zfill(11)
    return digits.zfill(14)


def _xlsx_visualization_path(csv_path: Path) -> Path:
    normalized_csv_path = csv_path
    if csv_path.suffix.lower() == ".tmp" and csv_path.name.lower().endswith(".csv.tmp"):
        normalized_csv_path = csv_path.with_suffix("")
    return normalized_csv_path.with_name(f"{normalized_csv_path.stem}_excel.xlsx")


def _save_excel_visualization(df: pd.DataFrame, csv_path: Path) -> Path:
    xlsx_path = _xlsx_visualization_path(csv_path)
    tmp_xlsx = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")

    try:
        wb = Workbook()
        ws = wb.active

        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=str(col_name))

        for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell_value = "" if value is None else str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if col_idx == 3:
                    cell.number_format = "@"

        wb.save(tmp_xlsx)
        os.replace(tmp_xlsx, xlsx_path)
    finally:
        if tmp_xlsx.exists():
            try:
                tmp_xlsx.unlink()
            except Exception:
                pass
    return xlsx_path


class FTP_TLS_Implicit(FTP_TLS):
    def connect(self, host="", port=0, timeout=-999):
        if timeout != -999:
            self.timeout = timeout
        if port > 0:
            self.port = port
        self.host = host
        sock = socket.create_connection((self.host, self.port), self.timeout)
        self.af = sock.family
        self.sock = sock
        self.file = None
        if self.context is None:
            self.context = ssl.create_default_context()
        self.sock = self.context.wrap_socket(self.sock, server_hostname=self.host)
        self.file = self.sock.makefile("r", encoding=self.encoding)
        self.welcome = self.getresp()
        return self.welcome


def _criar_contexto_tls():
    contexto = ssl.create_default_context()
    if not FTP_TLS_VERIFY:
        contexto.check_hostname = False
        contexto.verify_mode = ssl.CERT_NONE
    if FTP_TLS_LEGACY:
        try:
            contexto.minimum_version = ssl.TLSVersion.TLSv1
        except Exception:
            pass
        try:
            contexto.set_ciphers("DEFAULT:@SECLEVEL=1")
        except Exception:
            pass
    return contexto


def _abrir_conexao_ftp(contexto, passive=True, data_protection="P", ftp_dir: str | None = None):
    if FTP_USE_TLS and FTP_TLS_IMPLICIT:
        ftp = FTP_TLS_Implicit()
    elif FTP_USE_TLS:
        ftp = FTP_TLS()
    else:
        ftp = FTP()

    if contexto is not None:
        ftp.context = contexto
    logger.info("Conectando ao FTP %s:%s (timeout=%ss)", FTP_HOST, FTP_PORT, FTP_TIMEOUT)
    ftp.connect(FTP_HOST, FTP_PORT, timeout=FTP_TIMEOUT)
    if ftp.sock:
        ftp.sock.settimeout(FTP_TIMEOUT)

    logger.info("Autenticando no FTP")
    ftp.login(FTP_USER, FTP_PASSWORD)

    if FTP_USE_TLS:
        if data_protection.upper() == "C":
            logger.info("Canal de dados sem TLS (prot_c)")
            ftp.prot_c()
        else:
            logger.info("Canal de dados com TLS (prot_p)")
            ftp.prot_p()

    ftp.set_pasv(passive)
    logger.info("Modo de transferencia FTP: %s", "PASV" if passive else "ATIVO")

    destino_ftp = str(ftp_dir or FTP_DIR)
    logger.info("Alterando diretorio no FTP: %s", destino_ftp)
    ftp.cwd(destino_ftp)
    return ftp

def abrir_pasta_exportacao(driver, wait):
    del driver, wait
    logger.info(
        "Compatibilidade legado: abrir_pasta_exportacao nao navega mais pela UI; a pasta '%s' sera consultada no navegador autenticado.",
        FILE_MANAGER_EXPORT_FOLDER,
    )


def _mapear_csvs_legado(itens):
    csvs = []
    for item in itens:
        nome = item.get("name")
        if not isinstance(nome, str):
            continue
        if bool(item.get("isDirectory")):
            continue
        if not nome.lower().endswith(".csv"):
            continue
        csvs.append(
            {
                "nome": nome,
                "data_modificacao": item.get("dateModified") or "",
            }
        )
    return csvs


def listar_csvs(driver, wait):
    del wait
    try:
        itens = listar_arquivos_api_no_browser(driver, FILE_MANAGER_EXPORT_FOLDER)
        logger.info(
            "Listagem via navegador concluida na pasta '%s': %s item(ns)",
            FILE_MANAGER_EXPORT_FOLDER,
            len(itens),
        )
        csvs = _mapear_csvs_legado(itens)
        if not csvs:
            raise RuntimeError("Nenhum arquivo CSV encontrado no navegador autenticado.")
        logger.info("Total de CSVs encontrados no navegador: %s", len(csvs))
        return csvs
    except Exception as exc:
        logger.exception(
            "Falha ao listar arquivos no navegador autenticado na pasta '%s'",
            FILE_MANAGER_EXPORT_FOLDER,
        )
        raise RuntimeError(
            f"Falha ao listar arquivos na pasta '{FILE_MANAGER_EXPORT_FOLDER}'."
        ) from exc


def _parse_data(data_str):
    formatos = ["%d/%m/%Y", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"]
    for fmt in formatos:
        try:
            return datetime.strptime(data_str, fmt)
        except ValueError:
            continue
    raise RuntimeError(f"Formato de data invalido: {data_str}")


def _parse_data_nome(nome):
    match = re.search(r"(\d{8})", nome)
    if not match:
        return None
    return datetime.strptime(match.group(1), "%Y%m%d")


def selecionar_csv_mais_recente(csvs):
    for item in csvs:
        data_mod = item.get("data_modificacao", "")
        try:
            item["data_parseada"] = _parse_data(data_mod)
        except Exception:
            data_nome = _parse_data_nome(item.get("nome", ""))
            if not data_nome:
                raise RuntimeError("Nao foi possivel extrair a data do arquivo.")
            item["data_parseada"] = data_nome
    ordenados = sorted(csvs, key=lambda x: x["data_parseada"], reverse=True)
    logger.info("CSV mais recente selecionado: %s", ordenados[0].get("nome"))
    return ordenados[0]


def baixar_arquivo(driver, wait, nome_arquivo):
    try:
        logger.info(
            "Iniciando download via form submit no navegador do arquivo '%s' na pasta '%s'",
            nome_arquivo,
            FILE_MANAGER_EXPORT_FOLDER,
        )
        baixar_arquivo_via_form_submit_no_browser(
            driver,
            FILE_MANAGER_EXPORT_FOLDER,
            nome_arquivo,
        )
        logger.info("Download disparado. Aguardando arquivo em %s", DOWNLOAD_DIR)
        baixado = aguardar_download(wait, nome_arquivo)
        if not baixado.exists():
            raise RuntimeError(
                f"Download nao encontrado em disco apos disparo no navegador: {baixado}"
            )
        if int(baixado.stat().st_size) <= 0:
            raise RuntimeError(f"Arquivo baixado esta vazio: {baixado.name}")
        logger.info("Download via navegador concluido para: %s", nome_arquivo)
        return baixado
    except Exception as exc:
        logger.exception("Falha ao baixar arquivo via form submit no navegador: %s", nome_arquivo)
        raise RuntimeError(
            f"Falha ao baixar o arquivo '{nome_arquivo}' no navegador."
        ) from exc


def limpar_arquivos_download_anteriores(nome_arquivo: str) -> None:
    download_dir = Path(DOWNLOAD_DIR)
    if not download_dir.exists():
        logger.info("Limpeza pre-download: diretorio inexistente (%s).", download_dir)
        return
    stem = Path(nome_arquivo).stem
    suffix = Path(nome_arquivo).suffix
    var_regex = re.compile(rf"^{re.escape(stem)} \(\d+\){re.escape(suffix)}$")
    var_temp_regex = re.compile(
        rf"^{re.escape(stem)} \(\d+\){re.escape(suffix)}\.crdownload$"
    )

    removidos = []
    for item in download_dir.iterdir():
        if not item.is_file():
            continue
        nome = item.name
        is_target = (
            nome == nome_arquivo
            or nome == f"{nome_arquivo}.crdownload"
            or bool(var_regex.match(nome))
            or bool(var_temp_regex.match(nome))
        )
        if not is_target:
            continue
        try:
            item.unlink()
            removidos.append(nome)
        except Exception as exc:
            logger.warning("Falha ao remover arquivo antigo de download (%s): %s", nome, exc)

    if removidos:
        logger.info(
            "Limpeza pre-download removeu %s arquivo(s): %s",
            len(removidos),
            ", ".join(sorted(removidos)),
        )
    else:
        logger.info("Limpeza pre-download: nenhum arquivo antigo para %s", nome_arquivo)


def encontrar_download_real(nome_arquivo: str) -> Path | None:
    download_dir = Path(DOWNLOAD_DIR)
    if not download_dir.exists():
        return None
    esperado = download_dir / nome_arquivo
    if esperado.exists():
        return esperado

    stem = Path(nome_arquivo).stem
    suffix = Path(nome_arquivo).suffix
    var_regex = re.compile(rf"^{re.escape(stem)} \(\d+\){re.escape(suffix)}$")

    candidatos = []
    for item in download_dir.iterdir():
        if not item.is_file():
            continue
        if not var_regex.match(item.name):
            continue
        try:
            if item.stat().st_size <= 0:
                continue
        except OSError:
            continue
        candidatos.append(item)

    if not candidatos:
        return None
    return max(candidatos, key=lambda p: p.stat().st_mtime)


def aguardar_download(wait, nome_arquivo):
    del wait
    deadline = datetime.now().timestamp() + float(DOWNLOAD_WAIT_TIMEOUT_SECONDS)
    arquivo_final = None

    while datetime.now().timestamp() < deadline:
        arquivo_final = Path(DOWNLOAD_DIR) / nome_arquivo
        if not arquivo_final.exists():
            arquivo_final = encontrar_download_real(nome_arquivo)

        arquivo_temporario = Path(DOWNLOAD_DIR) / f"{nome_arquivo}.crdownload"
        if (
            arquivo_final is not None
            and arquivo_final.exists()
            and not arquivo_temporario.exists()
        ):
            try:
                if int(arquivo_final.stat().st_size) > 0:
                    logger.info("Validacao do download concluida: %s", arquivo_final.name)
                    return arquivo_final
            except OSError:
                pass

        sleep(float(DOWNLOAD_WAIT_POLL_SECONDS))

    raise RuntimeError(
        f"Download nao apareceu no diretorio {DOWNLOAD_DIR} para o arquivo esperado '{nome_arquivo}'."
    )


def renomear_arquivo_baixado(nome_arquivo):
    download_dir = Path(DOWNLOAD_DIR)
    origem = download_dir / nome_arquivo
    if not origem.exists():
        raise RuntimeError(f"Arquivo nao encontrado para renomear: {nome_arquivo}")

    sufixo_data = datetime.now().strftime("%Y%m%d_%H%M%S")
    novo_nome = f"{FILE_PREFIX}{sufixo_data}.csv"
    destino = download_dir / novo_nome

    if destino.exists():
        raise RuntimeError(f"Arquivo de destino ja existe: {novo_nome}")

    origem.rename(destino)
    logger.info("Arquivo renomeado: %s -> %s", nome_arquivo, destino.name)
    return destino


def remover_cabecalho_csv(caminho_arquivo):
    caminho = Path(caminho_arquivo)
    if not caminho.exists():
        raise RuntimeError(f"Arquivo nao encontrado para limpar cabecalho: {caminho}")

    for tentativa in range(5):
        try:
            tmp_path = caminho.with_suffix(caminho.suffix + ".tmp")
            raw_content = caminho.read_bytes()
            try:
                raw_content.decode("utf-8")
            except UnicodeDecodeError:
                normalized_text = raw_content.decode("cp1252")
                caminho.write_bytes(normalized_text.encode("utf-8"))
                logger.warning(
                    "CSV convertido de cp1252 para UTF-8 antes do tratamento: %s",
                    caminho.name,
                )

            df = pd.read_csv(
                caminho,
                sep=";",
                encoding="utf-8",
                dtype=str,
                keep_default_na=False,
            )

            if df.shape[1] < 3:
                raise ValueError("Arquivo nao possui coluna C suficiente para tratamento.")

            df = df.fillna("")
            df.iloc[:, 2] = df.iloc[:, 2].map(_normalize_cpf_cnpj)

            df.to_csv(
                tmp_path,
                sep=";",
                header=False,
                index=False,
                encoding="utf-8",
            )
            os.replace(tmp_path, caminho)
            xlsx_path = _save_excel_visualization(df, caminho)
            logger.info("Cabecalho removido do CSV: %s", caminho.name)
            logger.info("Arquivo XLSX de visualizacao gerado: %s", xlsx_path.name)
            return
        except OSError as exc:
            try:
                if "tmp_path" in locals() and tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass
            if tentativa >= 4:
                raise
            logger.warning("Falha ao remover cabecalho, tentando novamente: %s", exc)
            sleep(1.0)


def copiar_arquivo(caminho_arquivo, copy_dir: str | None = None):
    origem = Path(caminho_arquivo)
    if not origem.exists():
        raise RuntimeError(f"Arquivo nao encontrado para copiar: {origem}")

    destino_dir = Path(copy_dir or COPY_DIR)
    destino_dir.mkdir(parents=True, exist_ok=True)
    destino = destino_dir / origem.name
    shutil.copy2(origem, destino)
    logger.info("Arquivo copiado para servidor: %s", destino)
    return destino


def _verificar_upload_ftp(contexto, nome_arquivo, tamanho_local, ftp_dir: str | None = None):
    ftp = None
    try:
        ftp = _abrir_conexao_ftp(contexto, passive=True, data_protection="P", ftp_dir=ftp_dir)
        try:
            tamanho_remoto = ftp.size(nome_arquivo)
        except Exception:
            tamanho_remoto = None

        if tamanho_remoto is None:
            try:
                nomes = ftp.nlst()
                existe = nome_arquivo in nomes
            except Exception:
                existe = False
            logger.info(
                "Verificacao FTP (sem SIZE): existe=%s arquivo=%s",
                existe,
                nome_arquivo,
            )
            return False

        ok = int(tamanho_remoto) == int(tamanho_local)
        logger.info(
            "Verificacao FTP: arquivo=%s tamanho_local=%s tamanho_remoto=%s ok=%s",
            nome_arquivo,
            tamanho_local,
            tamanho_remoto,
            ok,
        )
        return ok
    except Exception:
        logger.exception("Falha ao verificar arquivo no FTP: %s", nome_arquivo)
        return False
    finally:
        if ftp:
            try:
                ftp.quit()
            except Exception:
                try:
                    ftp.close()
                except Exception:
                    pass


def enviar_arquivo_ftp(caminho_arquivo, ftp_dir: str | None = None):
    origem = Path(caminho_arquivo)
    if not origem.exists():
        raise RuntimeError(f"Arquivo nao encontrado para enviar ao FTP: {origem}")

    logger.info("Iniciando envio ao FTP: %s", origem.name)
    tamanho_local = origem.stat().st_size
    contexto = None
    if FTP_USE_TLS:
        contexto = _criar_contexto_tls()

    protecoes = [FTP_DATA_PROTECTION.upper()]
    if protecoes[0] == "P":
        protecoes.append("C")
    else:
        protecoes.append("P")

    passivos = [True] if FTP_PASSIVE else [False]
    for passive in passivos:
        for protecao in protecoes:
            ftp = None
            modo = "PASV" if passive else "ATIVO"
            try:
                ftp = _abrir_conexao_ftp(
                    contexto,
                    passive=passive,
                    data_protection=protecao,
                    ftp_dir=ftp_dir,
                )
                logger.info(
                    "Enviando arquivo via STOR: %s (protecao=%s, modo=%s)",
                    origem.name,
                    protecao,
                    modo,
                )
                with origem.open("rb") as arquivo:
                    ftp.storbinary(f"STOR {origem.name}", arquivo, blocksize=64 * 1024)
                logger.info("Arquivo enviado ao FTP: %s", origem.name)
                return
            except error_perm as exc:
                if "common algorithm" in str(exc).lower():
                    logger.warning(
                        "Falha de algoritmo TLS no envio (protecao=%s, modo=%s).",
                        protecao,
                        modo,
                    )
                elif "policy requires ssl" in str(exc).lower():
                    logger.warning(
                        "Servidor exige SSL, ignorando protecao=%s (modo=%s).",
                        protecao,
                        modo,
                    )
                else:
                    logger.exception("Falha no envio ao FTP")
                    raise
            except TimeoutError:
                logger.warning(
                    "Timeout no envio/fechamento TLS (protecao=%s, modo=%s).",
                    protecao,
                    modo,
                )
                if FTP_USE_TLS and _verificar_upload_ftp(
                    contexto,
                    origem.name,
                    tamanho_local,
                    ftp_dir=ftp_dir,
                ):
                    logger.warning(
                        "Arquivo parece ter sido enviado com sucesso apesar do timeout (verificado por SIZE): %s",
                        origem.name,
                    )
                    return
            except ConnectionResetError:
                logger.warning(
                    "Conexao resetada no envio (protecao=%s, modo=%s).",
                    protecao,
                    modo,
                )
                if FTP_USE_TLS and _verificar_upload_ftp(
                    contexto,
                    origem.name,
                    tamanho_local,
                    ftp_dir=ftp_dir,
                ):
                    logger.warning(
                        "Arquivo parece ter sido enviado com sucesso apesar do reset (verificado por SIZE): %s",
                        origem.name,
                    )
                    return
            except Exception:
                logger.exception("Falha no envio ao FTP")
                raise
            finally:
                if ftp:
                    try:
                        ftp.quit()
                    except Exception:
                        try:
                            ftp.close()
                        except Exception:
                            pass

    raise RuntimeError("Falha no envio ao FTP em todas as tentativas.")


def arquivo_ftp_existe_com_mesmo_tamanho(caminho_arquivo, ftp_dir: str | None = None):
    origem = Path(caminho_arquivo)
    if not origem.exists():
        raise RuntimeError(f"Arquivo nao encontrado para verificar no FTP: {origem}")

    tamanho_local = origem.stat().st_size
    contexto = None
    if FTP_USE_TLS:
        contexto = _criar_contexto_tls()

    try:
        return _verificar_upload_ftp(
            contexto,
            origem.name,
            tamanho_local,
            ftp_dir=ftp_dir,
        )
    except Exception:
        logger.exception("Falha ao verificar existencia do arquivo no FTP: %s", origem.name)
        return False


def testar_conexao_ftp():
    logger.info("Testando conexao com FTP")
    contexto = None
    ftp = None
    try:
        if FTP_USE_TLS:
            contexto = _criar_contexto_tls()

        ftp = _abrir_conexao_ftp(contexto, passive=True)
        logger.info("Conexao FTP validada com sucesso")
        return True
    except Exception:
        logger.exception("Falha ao testar conexao FTP")
        return False
    finally:
        if ftp:
            try:
                ftp.quit()
            except Exception:
                try:
                    ftp.close()
                except Exception:
                    pass
