import unittest
from pathlib import Path

import main
import pandas as pd
from openpyxl import load_workbook
from tests._workspace_temp import workspace_temp_dir


class CsvProcessingTests(unittest.TestCase):
    def test_cp1252_input_is_normalized_and_processed(self) -> None:
        with workspace_temp_dir("csv_processing") as tmpdir:
            csv_path = Path(tmpdir) / "legacy_encoding.csv"
            legacy_content = (
                "col1;col2;col3;col4\n"
                "linha_descartada;xx;999;yy\n"
                "001;a\xe7\xe3o;000045;teste\n"
            ).encode("cp1252")
            csv_path.write_bytes(legacy_content)

            main.remover_cabecalho_csv(csv_path)

            raw = csv_path.read_bytes()
            decoded = raw.decode("utf-8")
            self.assertNotIn("col1;col2;col3;col4", decoded)

            df = pd.read_csv(
                csv_path,
                sep=";",
                encoding="utf-8",
                header=None,
                names=["col1", "col2", "col3", "col4"],
                dtype=str,
                keep_default_na=False,
            )
            self.assertEqual(len(df), 2)
            self.assertEqual(df.iloc[0, 2], "00000000999")
            self.assertEqual(df.iloc[1, 2], "00000000045")
            self.assertFalse(df.isna().any().any())

    def test_remove_header_and_preserve_column_c_as_text(self) -> None:
        with workspace_temp_dir("csv_processing") as tmpdir:
            csv_path = Path(tmpdir) / "sample.csv"
            csv_path.write_text(
                "col1;col2;col3;col4\n"
                "linha_descartada;xx;999;yy\n"
                "001;abc;000045;acao\n"
                "002;def;;teste\n",
                encoding="utf-8",
            )

            main.remover_cabecalho_csv(csv_path)

            df = pd.read_csv(
                csv_path,
                sep=";",
                encoding="utf-8",
                header=None,
                names=["col1", "col2", "col3", "col4"],
                dtype=str,
                keep_default_na=False,
            )

            self.assertEqual(len(df), 3)
            self.assertEqual(df.iloc[0, 2], "00000000999")
            self.assertEqual(df.iloc[1, 2], "00000000045")
            self.assertEqual(df.iloc[2, 2], "")
            self.assertFalse(df.isna().any().any())

    def test_normalize_column_c_to_digits_only(self) -> None:
        with workspace_temp_dir("csv_processing") as tmpdir:
            csv_path = Path(tmpdir) / "normalize.csv"
            csv_path.write_text(
                "col1;col2;col3;col4\n"
                "linha_descartada;xx;999;yy\n"
                "1;a;123.456.789-01;ok\n"
                "2;b;12.345.678/0001-95;ok\n"
                "3;c;6,01433E+13;ok\n"
                "4;d;6.01433E+13;ok\n"
                "5;e;3,571E+28;ok\n",
                encoding="utf-8",
            )

            main.remover_cabecalho_csv(csv_path)

            df = pd.read_csv(
                csv_path,
                sep=";",
                encoding="utf-8",
                header=None,
                names=["col1", "col2", "col3", "col4"],
                dtype=str,
                keep_default_na=False,
            )

            self.assertEqual(len(df), 6)
            self.assertEqual(df.iloc[0, 2], "00000000999")
            self.assertEqual(df.iloc[1, 2], "12345678901")
            self.assertEqual(df.iloc[2, 2], "12345678000195")
            self.assertEqual(df.iloc[3, 2], "60143300000000")
            self.assertEqual(df.iloc[4, 2], "60143300000000")
            self.assertEqual(df.iloc[5, 2], "35710000000000000000000000000")

    def test_generate_xlsx_visualization_with_column_c_as_text(self) -> None:
        with workspace_temp_dir("csv_processing") as tmpdir:
            csv_tmp_path = Path(tmpdir) / "Exportacao_123.csv.tmp"
            csv_tmp_path.write_text(
                "col1;col2;col3;col4\n"
                "linha_descartada;xx;999;yy\n"
                "10;abc;00123456789;ok\n"
                "11;def;6,01433E+13;ok\n",
                encoding="utf-8",
            )

            main.remover_cabecalho_csv(csv_tmp_path)

            xlsx_path = Path(tmpdir) / "Exportacao_123_excel.xlsx"
            self.assertTrue(xlsx_path.exists())

            wb = load_workbook(xlsx_path)
            ws = wb.active

            self.assertEqual(ws["A1"].value, "col1")
            self.assertEqual(ws["C2"].value, "00000000999")
            self.assertEqual(ws["C3"].value, "00123456789")
            self.assertEqual(ws["C4"].value, "60143300000000")
            self.assertEqual(ws["C2"].number_format, "@")
            self.assertEqual(ws["C3"].number_format, "@")
            self.assertEqual(ws["C4"].number_format, "@")

    def test_raise_when_file_has_less_than_three_columns(self) -> None:
        with workspace_temp_dir("csv_processing") as tmpdir:
            csv_path = Path(tmpdir) / "invalid.csv"
            csv_path.write_text(
                "col1;col2\n"
                "linha_descartada;xx\n"
                "001;abc\n",
                encoding="utf-8",
            )

            with self.assertRaisesRegex(
                ValueError, "Arquivo nao possui coluna C suficiente para tratamento."
            ):
                main.remover_cabecalho_csv(csv_path)


if __name__ == "__main__":
    unittest.main()
